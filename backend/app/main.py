from __future__ import annotations

import csv
import io
import logging
import uuid

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.config import ALLOWED_ORIGINS, ALLOWED_CONTENT_TYPES, MAX_UPLOAD_MB, resolve_provider
from app.models import JobInfo, MTOResponse
from app.pipeline.pipeline import run_pipeline

logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Isometric to MTO API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory job store for synchronous processing and exports.
_jobs: dict[str, JobInfo] = {}


async def _read_and_validate(file: UploadFile) -> bytes:
    if file.content_type not in ALLOWED_CONTENT_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{file.content_type}'. Allowed: PNG, JPG, PDF.",
        )

    contents = await file.read()
    size_mb = len(contents) / (1024 * 1024)
    if size_mb > MAX_UPLOAD_MB:
        raise HTTPException(
            status_code=400,
            detail=f"File too large ({size_mb:.1f} MB). Max allowed is {MAX_UPLOAD_MB} MB.",
        )
    if len(contents) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    return contents


def _get_job(job_id: str) -> JobInfo:
    job = _jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="job_id not found")
    return job


def _create_processing_job() -> JobInfo:
    job_id = uuid.uuid4().hex
    job = JobInfo(job_id=job_id, status="processing")
    _jobs[job_id] = job
    return job


def _finalize_job(job: JobInfo, result: MTOResponse) -> None:
    result.result_id = job.job_id
    job.status = "complete"
    job.mto = result
    _jobs[job.job_id] = job


def _fail_job(job: JobInfo, exc: Exception) -> None:
    job.status = "failed"
    job.detail = str(exc)
    _jobs[job.job_id] = job


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/")
def root():
    return {
        "status": "ok",
        "message": "Isometric to MTO API is running. Use /api/health and /api/upload.",
        "provider": resolve_provider(),
    }


@app.post("/api/upload", response_model=JobInfo)
async def upload(file: UploadFile = File(...)):
    contents = await _read_and_validate(file)
    job = _create_processing_job()

    try:
        result = run_pipeline(contents, file.content_type, file.filename or "upload")
        _finalize_job(job, result)
    except Exception as exc:  # noqa: BLE001
        logging.exception("Pipeline failed unexpectedly")
        _fail_job(job, exc)
        raise HTTPException(
            status_code=500,
            detail=f"Extraction pipeline failed: {exc}",
        ) from exc

    return job


@app.get("/api/mto/{job_id}", response_model=JobInfo)
def get_mto(job_id: str):
    return _get_job(job_id)


@app.get("/api/mto/{job_id}/csv")
def get_mto_csv(job_id: str):
    job = _get_job(job_id)
    if job.status == "failed":
        raise HTTPException(status_code=409, detail=f"Job failed: {job.detail}")
    if job.status != "complete" or job.mto is None:
        raise HTTPException(status_code=202, detail="Job is still processing.")

    return _stream_csv(job.mto)


@app.get("/api/mto/{job_id}/xlsx")
def get_mto_xlsx(job_id: str):
    job = _get_job(job_id)
    if job.status == "failed":
        raise HTTPException(status_code=409, detail=f"Job failed: {job.detail}")
    if job.status != "complete" or job.mto is None:
        raise HTTPException(status_code=202, detail="Job is still processing.")

    target = job.mto
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    title_fill = PatternFill("solid", fgColor="1E293B")
    title_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    bold_font = Font(bold=True)

    summary_sheet["A1"] = "Isometric MTO Export"
    summary_sheet["A1"].font = title_font
    summary_sheet["A1"].fill = title_fill
    summary_sheet["A1"].alignment = Alignment(horizontal="center")
    summary_sheet.merge_cells("A1:B1")

    summary_sheet["A3"] = "Field"
    summary_sheet["B3"] = "Value"
    for cell in summary_sheet[3]:
        cell.font = bold_font
        cell.fill = header_fill

    meta_rows = [
        ("Drawing No", target.drawing_meta.drawing_no),
        ("Revision", target.drawing_meta.revision),
        ("Line Number", target.drawing_meta.line_number),
        ("NPS", target.drawing_meta.nps),
        ("Material Class", target.drawing_meta.material_class),
        ("Service", target.drawing_meta.service),
        ("Total Pipe Length (m)", round(target.summary.total_pipe_length_m, 2)),
        ("Fittings", target.summary.fittings),
        ("Flanges", target.summary.flanges),
        ("Valves", target.summary.valves),
        ("Gaskets", target.summary.gaskets),
        ("Bolt Sets", target.summary.bolt_sets),
        ("Supports", target.summary.supports),
        ("Instrumentation Connections", target.summary.instrumentation_connections),
        ("Field Welds", target.summary.field_welds),
    ]

    row_idx = 4
    for label, value in meta_rows:
        summary_sheet.cell(row=row_idx, column=1, value=label)
        summary_sheet.cell(row=row_idx, column=2, value=value if value is not None else "—")
        row_idx += 1

    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 24

    items_sheet = workbook.create_sheet("MTO_Items")
    headers = [
        "Item No", "Category", "Description", "Size", "Sched/Class", "Material",
        "End", "Qty", "Unit", "Length (m)", "Confidence", "Verification", "Derived From", "Pg", "Remarks",
    ]
    items_sheet.append(headers)
    for cell in items_sheet[1]:
        cell.font = bold_font
        cell.fill = header_fill

    for item in target.items:
        items_sheet.append([
            item.item_no,
            item.category,
            item.description,
            item.size_nps,
            item.schedule_rating,
            item.material_spec,
            item.end_type,
            item.quantity,
            item.unit,
            item.length_m,
            item.confidence,
            item.verification_status,
            item.derived_from,
            item.source_page,
            item.remarks,
        ])

    items_sheet.freeze_panes = "A2"
    items_sheet.auto_filter.ref = items_sheet.dimensions
    for column_cells in items_sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        items_sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mto_export.xlsx"},
    )


def _stream_csv(target: MTOResponse) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "item_no", "category", "description", "size_nps", "schedule_rating",
            "material_spec", "end_type", "quantity", "unit", "length_m",
            "confidence", "verification_status", "derived_from", "source_page", "remarks",
        ]
    )
    for item in target.items:
        writer.writerow(
            [
                item.item_no,
                item.category,
                item.description,
                item.size_nps,
                item.schedule_rating,
                item.material_spec,
                item.end_type,
                item.quantity,
                item.unit,
                item.length_m,
                item.confidence,
                item.verification_status,
                item.derived_from,
                item.source_page,
                item.remarks,
            ]
        )

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=mto_export.csv"},
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(request, exc):
    logging.exception("Unhandled exception")
    return JSONResponse(
        status_code=500,
        content={"error": "internal_error", "detail": str(exc)},
    )
