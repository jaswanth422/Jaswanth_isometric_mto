import io

import fitz
from fastapi.testclient import TestClient
from PIL import Image
from openpyxl import load_workbook

from app.main import app
from app.pipeline.pipeline import _pdf_to_image_bytes

client = TestClient(app)


def _sample_png_bytes() -> bytes:
    img = Image.new("RGB", (100, 100), color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _sample_pdf_bytes() -> bytes:
    document = fitz.open()
    page = document.new_page(width=100, height=100)
    page.insert_text((10, 50), "Test PDF", fontsize=12)
    return document.tobytes()


def test_health():
    resp = client.get("/api/health")
    assert resp.status_code == 200
    assert resp.json() == {"status": "ok"}


def test_upload_mock_happy_path(monkeypatch):
    monkeypatch.setenv("MODEL_PROVIDER", "mock")

    files = {"file": ("test.png", _sample_png_bytes(), "image/png")}
    resp = client.post("/api/upload", files=files)

    assert resp.status_code == 200
    job = resp.json()
    assert job["status"] == "complete"
    assert job["job_id"]
    assert job["mto"]["mock"] is True
    assert job["mto"]["mock_reason"] == "configured_mock_provider"
    assert len(job["mto"]["items"]) > 0

    categories = [item["category"] for item in job["mto"]["items"]]
    assert "GASKET" in categories
    assert "BOLT" in categories
    assert "SUPPORT" in categories
    assert "INSTRUMENTATION" in categories


def test_upload_rejects_bad_content_type():
    files = {"file": ("test.txt", b"not an image", "text/plain")}
    resp = client.post("/api/upload", files=files)
    assert resp.status_code == 400


def test_upload_rejects_empty_file():
    files = {"file": ("test.png", b"", "image/png")}
    resp = client.post("/api/upload", files=files)
    assert resp.status_code == 400


def test_pdf_render_falls_back_without_poppler(monkeypatch):
    import pdf2image

    def _raise_poppler_error(*args, **kwargs):
        raise RuntimeError("Unable to get page count. Is poppler installed and in PATH?")

    monkeypatch.setattr(pdf2image, "convert_from_bytes", _raise_poppler_error)

    image_bytes = _pdf_to_image_bytes(_sample_pdf_bytes())
    assert image_bytes[:8] == b"\x89PNG\r\n\x1a\n"


def test_get_mto_job_returns_job_info(monkeypatch):
    monkeypatch.setenv("MODEL_PROVIDER", "mock")
    files = {"file": ("test.png", _sample_png_bytes(), "image/png")}
    resp = client.post("/api/upload", files=files)
    assert resp.status_code == 200

    job_id = resp.json()["job_id"]
    job_resp = client.get(f"/api/mto/{job_id}")

    assert job_resp.status_code == 200
    assert job_resp.json()["job_id"] == job_id
    assert job_resp.json()["status"] == "complete"


def test_mto_csv_returns_csv(monkeypatch):
    monkeypatch.setenv("MODEL_PROVIDER", "mock")
    files = {"file": ("test.png", _sample_png_bytes(), "image/png")}
    resp = client.post("/api/upload", files=files)
    assert resp.status_code == 200

    job_id = resp.json()["job_id"]
    csv_resp = client.get(f"/api/mto/{job_id}/csv")

    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    assert "item_no,category,description" in csv_resp.text


def test_csv_requires_prior_upload():
    resp = client.get("/api/mto/doesnotexist/csv")
    assert resp.status_code == 404


def test_mto_xlsx_returns_workbook(monkeypatch):
    monkeypatch.setenv("MODEL_PROVIDER", "mock")

    files = {"file": ("test.png", _sample_png_bytes(), "image/png")}
    resp = client.post("/api/upload", files=files)
    assert resp.status_code == 200

    job_id = resp.json()["job_id"]
    xlsx_resp = client.get(f"/api/mto/{job_id}/xlsx")
    assert xlsx_resp.status_code == 200
    assert xlsx_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(io.BytesIO(xlsx_resp.content))
    assert workbook.sheetnames == ["Summary", "MTO_Items"]
    assert workbook["Summary"]["A1"].value == "Isometric MTO Export"
    assert workbook["Summary"]["A16"].value == "Supports"
    assert workbook["Summary"]["A17"].value == "Instrumentation Connections"


def test_field_weld_count_is_carried_into_summary(monkeypatch):
    from app.pipeline import pipeline as pipeline_module

    def fake_ocr(_image_bytes: bytes) -> dict:
        return {
            "has_bom": True,
            "drawing_no": None,
            "revision": None,
            "line_number": None,
            "field_welds": 3,
            "category_counts": {},
            "raw_text": "FW FW FIELD WELD",
        }

    monkeypatch.setattr(pipeline_module, "read_drawing", fake_ocr)
    monkeypatch.setattr(pipeline_module, "provider_context", lambda: ("mock", "configured_mock_provider"))

    files = {"file": ("test.png", _sample_png_bytes(), "image/png")}
    resp = client.post("/api/upload", files=files)

    assert resp.status_code == 200
    assert resp.json()["mto"]["summary"]["field_welds"] == 3


def test_multi_page_pdf_tracks_source_pages(monkeypatch):
    monkeypatch.setenv("MODEL_PROVIDER", "mock")

    document = fitz.open()
    first = document.new_page(width=100, height=100)
    first.insert_text((10, 50), "Page 1", fontsize=12)
    second = document.new_page(width=100, height=100)
    second.insert_text((10, 50), "Page 2", fontsize=12)

    resp = client.post(
        "/api/upload",
        files={"file": ("multi.pdf", document.tobytes(), "application/pdf")},
    )

    assert resp.status_code == 200
    data = resp.json()["mto"]
    source_pages = {item["source_page"] for item in data["items"]}
    assert source_pages == {1, 2}
    assert data["summary"]["supports"] >= 2
    assert data["summary"]["instrumentation_connections"] >= 2


def test_gemini_timeout_falls_back_once_and_opens_cooldown(monkeypatch):
    from app.pipeline import gemini_provider
    from app.pipeline import pipeline as pipeline_module

    calls = {"count": 0}

    def fake_gemini(_image_bytes: bytes, _mime_type: str, _filename: str) -> dict:
        calls["count"] += 1
        raise TimeoutError("simulated timeout")

    monkeypatch.setattr(pipeline_module, "provider_context", lambda: ("gemini", None))
    monkeypatch.setattr(gemini_provider, "extract_gemini", fake_gemini)
    monkeypatch.setattr(
        pipeline_module,
        "read_drawing",
        lambda _image_bytes: {
            "has_bom": False,
            "drawing_no": None,
            "revision": None,
            "line_number": None,
            "field_welds": 0,
            "supports": 0,
            "instrumentation": 0,
            "category_counts": {},
            "raw_text": "",
        },
    )
    monkeypatch.setattr(pipeline_module, "GEMINI_FAILURE_COOLDOWN_SECONDS", 60)
    pipeline_module._gemini_unavailable_until = 0.0

    first = pipeline_module.run_pipeline(_sample_png_bytes(), "image/png", "first.png")
    second = pipeline_module.run_pipeline(_sample_png_bytes(), "image/png", "second.png")

    assert first.mock is True
    assert first.provider == "mock"
    assert first.mock_reason == "gemini_error"
    assert "simulated timeout" in (first.mock_details or "")
    assert second.mock is True
    assert second.provider == "mock"
    assert second.mock_reason == "gemini_cooldown"
    assert "simulated timeout" in (second.mock_details or "")
    assert calls["count"] == 1
